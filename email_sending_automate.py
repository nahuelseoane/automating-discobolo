import smtplib
import pandas as pd
from filter_payments import extract_operation_number
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import os

# Email settings
SMTP_SERVER = '${SMTP_SERVER}'  # or your email provider's SMTP server
SMTP_PORT = ${SMTP_PORT}
EMAIL_USER = '${EMAIL_USER}'  # your email address
EMAIL_PASSWORD = '${EMAIL_PASSWORD}'  # your email password

# Define storage folders
main_file = "${BASE_PATH}/${YEAR}/Transferencias ${YEAR}.xlsx"
month = 'Marzo'
sheet_name = month
emails_file = "${BASE_PATH}/${YEAR}/EmailSocios.xlsx"
payments_path = f"${BASE_PATH}/${YEAR}/3 {month} ${YEAR}"
payments_folder = payments_path

# Load Excel files
df_main = pd.read_excel(main_file, sheet_name=sheet_name)
df_filtered = df_main[df_main["Concepto"].str.contains(
    "Cuota", case=False, na=False)].copy()
df_filtered.reset_index(drop=True, inplace=True)
print(
    f"✅ Loaded {len(df_filtered)} payments from {sheet_name} (Cuotas) only.")
df_emails = pd.read_excel(emails_file, sheet_name=sheet_name)


# Merging files
df_merged = df_filtered.merge(
    df_emails[['Nombre Completo', 'Emails']],
    left_on='Jefe de Grupo',  # Column in df_main
    right_on='Nombre Completo',  # Column in df_emails
    how='left'  # Keep all rows from df_main
)


def send_email(user, recipient_email, pdf_path):
    first_name = user.split(", ")[1].split()[0].lower().capitalize()
    if not os.path.exists(pdf_path):  # Check if PDF exists
        print(f"⚠️ Error: PDF not found for {user} ({pdf_path})")
        return False  # Return False if email was not sent

    msg = MIMEMultipart()
    msg['From'], msg['To'], msg['Subject'] = EMAIL_USER, recipient_email, 'Recibo de pago - Transferencia confirmada'

    # Email body text
    if first_name:
        body = f'Buenos días {first_name}:\n\nRecibimos su transferencia.\nAdjuntamos el recibo correspondiente.\n\nSaludos!'
        msg.attach(MIMEText(body, 'plain'))
    else:
        print(f"❌ Error extracting First Name from {user}.")
        return False

    # Attach PDF
    with open(pdf_path, 'rb') as file:
        part = MIMEApplication(file.read(), Name=os.path.basename(pdf_path))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(pdf_path)}"'
        msg.attach(part)

    # Send email
    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        server.send_message(msg)

    print(
        f'   ✅ Email successfully sent to {user} (First Name: {first_name}) (Email: {recipient_email}) ✅')
    return True  # Return True if email was successfully sent


# Loading workbook for later saving 'Si' in Email Sent column
wb_main = load_workbook(main_file)
if sheet_name not in wb_main.sheetnames:
    print(f"❌ {sheet_name} not find in Main Excel sheet names.")
ws_main = wb_main[sheet_name]

# Loop through each recipient and send the email
for index, row in df_merged.iterrows():
    # index_2 += 1
    transaction_number = extract_operation_number(row["Descripción"])
    user, email = row['Jefe de Grupo'], row['Emails']
    email = str(email) if isinstance(email, str) else ""
    email = email.split(";")[0].strip()
    if str(row['Email']).strip().lower() == "si":  # Skip emails already sent
        print(f"🔃 Skipping {user} - Email already sent.")
        continue

    if not email or "@" not in email:
        print(f"⚠️ Invalid or missing email for {user}. Skipping...")
        continue

    # Construct full PDF path
    pdf_filename = user.replace(",", "") + "_" + transaction_number + ".pdf"
    print(f"🔎 Sending email to {user}")
    pdf_path = os.path.join(payments_folder, pdf_filename)

    if send_email(user, email, pdf_path):  # If email was sent successfully
        # Mark email as sent in DataFrame
        for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row, values_only=False):
            if row[6].value == user:  # Assuming is Column G
                sent_cell = row[8]  # Column I (Sytech)
                sent_cell.value = "Si"  # ✅ Update the cell without affecting formatting

# ✅ Save the file without overwriting sheets or formatting
wb_main.save(main_file)
wb_main.close()

# df_main.to_excel(main_file, index=False)
print("🎉Emails sent successfully!")
