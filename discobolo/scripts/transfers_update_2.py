import pandas as pd
from openpyxl import load_workbook

from discobolo.config.config import (
    EMAILS_FILE,
    KEVIN_MAX,
    SHEET_NAME,
    TENNIS_CLASS_FEE,
    TRANSFER_FILE,
)
from discobolo.scripts.extra_functions import (
    ensure_sheet_exists,
    extract_deposito,
    extract_dni,
    filter_positive_payments,
)


def run_transfers_update_2():
    df, transfer = filter_positive_payments(TRANSFER_FILE, SHEET_NAME)
    ensure_sheet_exists(EMAILS_FILE, SHEET_NAME)
    df_emails = pd.read_excel(EMAILS_FILE, sheet_name=SHEET_NAME)

    transfer["DNI"] = transfer["Descripción"].apply(extract_dni)

    df_merged = transfer.merge(
        df_emails[["DNI", "Jefe de Grupo I", "Tipo de Pago"]], on="DNI", how="left"
    )

    wb = load_workbook(TRANSFER_FILE)
    ws = wb[SHEET_NAME]

    for idx, row in df_merged.iterrows():
        full_name = row["Jefe de Grupo I"]
        description = row["Descripción"]
        seq, amount, tipo = row["N° Secuencia"], row["Importe"], row["Tipo de Pago"]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            if row[0].value == seq:
                jefe_de_grupo = row[6]  # Column H
                concept = row[5]  # Column G
                if concept.value is None or concept.value == "":
                    if extract_deposito(description):
                        concept.value = "DEPOSITO"
                        print(f" ✅ {seq} Updated--> Concept: DEPOSITO")
                    elif not pd.isna(full_name):
                        if full_name == "CANFORA, KEVIN" and int(amount) <= KEVIN_MAX:
                            concept.value = "TENIS"
                            jefe_de_grupo.value = "CLASES KEVIN"
                            print(
                                f" ✅ {seq} Updated --> Concept: TENIS & Jefe de Grupo: CLASES KEVIN"
                            )
                            continue
                        elif not pd.isna(tipo):
                            concept.value = str(tipo).strip()
                        else:
                            concept.value = "CUOTA"

                        if jefe_de_grupo.value is None or jefe_de_grupo.value == "":
                            jefe_de_grupo.value = full_name
                            if full_name:
                                print(
                                    f" ✅ {seq} Updated --> Concept: {concept.value} & Jefe de Grupo: {full_name}."
                                )
                    elif int(amount) == TENNIS_CLASS_FEE:
                        concept.value = "TENIS"
                        jefe_de_grupo.value = "CLASE NO SOCIO"
                        print(
                            f" ✅ {seq} Updated --> Concept: TENIS & Jefe de Grupo: CLASE NO SOCIO"
                        )
                        continue
                    else:
                        print(f"  ❌ {seq} Not updated: Transfer's match wasn't found")

    wb.save(TRANSFER_FILE)
    wb.close()
    print("   ✅ Updated 'Jefe de Grupo' and 'Concepto'")


if __name__ == "__main__":
    run_transfers_update_2()
