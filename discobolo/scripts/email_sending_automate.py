import os
import smtplib
import pandas as pd
from datetime import datetime
from discobolo.scripts.extra_functions import extract_operation_number
from openpyxl import load_workbook
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from discobolo.config.config import (
    TRANSFER_FILE, SHEET_NAME, EMAILS_FILE, EMAIL_USER,
    SMTP_SERVER, SMTP_PORT, EMAIL_PASSWORD, PAYMENTS_PATH
)

def send_email(user, recipient_email, pdf_path):
    first_name = user.split(", ")[1].split()[0].lower().capitalize()
    if not os.path.exists(pdf_path):
        print(f"⚠️ Error: PDF not found for {user} ({pdf_path})")
        return False

    msg = MIMEMultipart()
    msg['From'], msg['To'], msg['Subject'] = EMAIL_USER, recipient_email, 'Recibo de pago - Transferencia confirmada'

    greetings = "Buenos días" if datetime.now().hour < 14 else "Buenas tardes"
    body = f'{greetings} {first_name}:\n\nRecibimos su transferencia.\nAdjuntamos el recibo correspondiente.\n\nSaludos!'
    msg.attach(MIMEText(body, 'plain'))

    with open(pdf_path, 'rb') as file:
        part = MIMEApplication(file.read(), Name=os.path.basename(pdf_path))
        part['Content-Disposition'] = f'attachment; filename="{os.path.basename(pdf_path)}"'
        msg.attach(part)

    with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
        server.login(EMAIL_USER, EMAIL_PASSWORD)
        server.send_message(msg)

    print(f'  ✅ Email sent to {user} (First Name: {first_name})')
    return True


def send_emails():
    # Load Excel files
    df_main = pd.read_excel(TRANSFER_FILE, sheet_name=SHEET_NAME)
    df_filtered = df_main[df_main["Concepto"].str.contains("Cuota", case=False, na=False)].copy()
    df_filtered.reset_index(drop=True, inplace=True)
    print(f"   🔃 Loaded {len(df_filtered)} payments from {SHEET_NAME} (Cuotas only).")

    df_emails = pd.read_excel(EMAILS_FILE, sheet_name=SHEET_NAME)

    df_merged = df_filtered.merge(
        df_emails[['Nombre Completo', 'Emails']],
        left_on='Jefe de Grupo',
        right_on='Nombre Completo',
        how='left'
    )

    wb_main = load_workbook(TRANSFER_FILE)
    if SHEET_NAME not in wb_main.sheetnames:
        print(f"❌ {SHEET_NAME} not found in Main Excel sheet names.")
        return
    ws_main = wb_main[SHEET_NAME]

    for index, row in df_merged.iterrows():
        user, email = row['Jefe de Grupo'], row['Emails']

        if pd.isna(row['Sytech']):
            print(f" ❌ Payment not yet uploaded - {user}")
            continue

        transaction_number = extract_operation_number(row["Descripción"])
        email = str(email) if isinstance(email, str) else ""
        email = email.split(";")[0].strip()

        if str(row.get('Email')).strip().lower() == "si":
            continue  # Already sent

        if not email or "@" not in email:
            print(f"⚠️ Invalid or missing email for {user}. Skipping...")
            continue

        pdf_filename = user.replace(",", "") + "_" + transaction_number + ".pdf"
        print(f"🔎 Sending email to {user}")
        pdf_path = os.path.join(PAYMENTS_PATH, pdf_filename)

        if send_email(user, email, pdf_path):
            for row in ws_main.iter_rows(min_row=2, max_row=ws_main.max_row, values_only=False):
                if row[6].value == user:
                    row[8].value = "Si"  # Column I
                    break

    wb_main.save(TRANSFER_FILE)
    wb_main.close()
    print("   🎉 Emails sent successfully!")
