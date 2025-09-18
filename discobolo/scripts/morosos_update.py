from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from discobolo.config.config import MONTH, MOROSOS_MAIN, MOROSOS_REPORT


def run_morosos_update():
    df_daily = pd.read_excel(MOROSOS_REPORT, skiprows=6, engine="openpyxl")
    wb = load_workbook(MOROSOS_MAIN)

    try:
        with pd.ExcelFile(MOROSOS_MAIN, engine="openpyxl") as xls:
            if MONTH not in xls.sheet_names:
                print(f"📂 {MONTH} sheet not found. Creating it...")
                wb.create_sheet(MONTH)
                wb.save(MOROSOS_MAIN)
                print(f"✅ {MONTH} sheet created successfully.")
    except FileNotFoundError:
        print("❌ Error: Main Morosos file not found!")

    ws = wb[MONTH]
    df_daily = df_daily.drop(
        columns=[
            "NROCUENTA",
            "FECHAULT.MOVDEBITO",
            "MONTOULT.MOVDEBITO",
            "DESC.ULT.MOVDEBITO",
            "FECHAULT.MOVCREDITO",
            "MONTOULT.MOVCREDITO",
            "TELEFONO",
            "EMAIL",
            "SUCURSALEMISORA",
        ]
    )
    df_daily = df_daily.sort_values(by=df_daily.columns[1], ascending=False)
    df_daily = df_daily.drop(index=0).iloc[1:].reset_index(drop=True)
    df_daily = df_daily[df_daily.iloc[:, 1] > 0]
    df_daily.columns.values[1] = "SALDO"

    with pd.ExcelWriter(
        MOROSOS_MAIN, engine="openpyxl", mode="a", if_sheet_exists="replace"
    ) as writer:
        df_daily.to_excel(writer, sheet_name=MONTH, index=False, startrow=1)
        worksheet = writer.sheets[MONTH]
        
        # Style in headers
        fill = PatternFill(start_color="4DA6D2", end_color="4DA6D2", fill_type="solid")
        font = Font(color="FFFFFF", bold=True)
        for cell in worksheet[2]:
            cell.fill = fill
            cell.font = font

        

        # Date in row 1
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        worksheet["A1"] = f"Última actualización: {now} hs."
        worksheet["A1"].font = Font(italic=True, color="888888")
        worksheet.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=len(df_daily.columns)
        )

        # Total at final row
        last_row = len(df_daily)
        row_total = last_row + 4
        last_column = ws.max_column

        print(f"last row: {last_row}")
        print(f"last column: {last_column}")

        worksheet[f"B{row_total}"] = f"=SUM(B3:B{last_row+2})"
        worksheet[f"B{row_total}"].font = Font(bold=True)
        worksheet[f"A{row_total}"] = "TOTAL"
        worksheet[f"A{row_total}"].font = Font(bold=True)

        # Column width
        for column_cells in worksheet.columns:
            max_len = max(
                (len(str(cell.value)) for cell in column_cells if cell.value), default=0
            )
            col_letter = get_column_letter(column_cells[0].column)
            worksheet.column_dimensions[col_letter].width = max_len + 2

        # 2nd column: currency format
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = '"$"#,##0.00'

    print("   ✅ Morosos main file updated.")
    wb.close()
