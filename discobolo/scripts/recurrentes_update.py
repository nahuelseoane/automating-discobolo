import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from discobolo.config.config import RECURRENTES_REPORT, RECURRENTES_MAIN, MONTH


df = pd.read_excel(RECURRENTES_REPORT, skiprows=4, engine="openpyxl")
wb = load_workbook(RECURRENTES_MAIN)
# Checking if actual month sheet exists
try:
    with pd.ExcelFile(RECURRENTES_MAIN, engine="openpyxl") as xls:
        sheet_names = xls.sheet_names  # Get all sheet names

    if MONTH not in sheet_names:
        print(f"📂 {MONTH} sheet not found. Creating it...")

        wb.create_sheet(MONTH)  # Create new sheet
        wb.save(RECURRENTES_MAIN)

        print(f"✅ {MONTH} sheet created successfully.")

except FileNotFoundError:
    print("❌ Error: Main Morosos file not found!")

# Function to categorize concepts
df.columns = df.columns.str.strip()

df = df[df["ACTIVO"].str.contains("S", case=False, na=False)]


def categorize(desc):
    if not isinstance(desc, str):
        return None
    if 'Tenis Mes' in desc:
        return 'Tenis Mes'
    elif 'Ropero' in desc:
        return 'Ropero'
    elif 'Cuota' in desc:
        return 'Cuotas'
    else:
        return None


# Add a new column to mark category
df['Category'] = df['DESC CONCEPTO'].apply(categorize)

# Step 3: Pivot the table
df = df.pivot_table(index="NOMBRE", columns="Category",
                          values="DESC CONCEPTO", aggfunc="first").reset_index()

# Optional: reorder columns
df = df[["NOMBRE", "Cuotas", "Tenis Mes", "Ropero"]]

# Step 4: Save to Excel or display
# df_pivot.to_excel("recurrentes_pivoted.xlsx", index=False)
# Pivot to reshape
# pivot = df.pivot_table(index='NOMBRE', columns='Category',
#                     values='DESC CONCEPTO', aggfunc='first').reset_index()

with pd.ExcelWriter(RECURRENTES_MAIN, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df.to_excel(writer, sheet_name=MONTH, index=False, startrow=0)
    # Access the workbook and the active worksheet
    workbook = writer.book
    worksheet = writer.sheets[MONTH]

    # ✔️ Adjusted styling
    header_fill = PatternFill(
        start_color='4DA6D2', end_color='4DA6D2', fill_type='solid')  # Excel-style blue
    header_font = Font(color='FFFFFF', bold=True)  # Bold white text

    # Apply the formatting to each header cell
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    # ✔️ Auto-adjust column widths
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max_length + 2


print("   ✅ 'Recurrentes' main file updated.")
wb.close()
