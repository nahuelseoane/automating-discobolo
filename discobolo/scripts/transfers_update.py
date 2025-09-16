from copy import copy
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from discobolo.config.config import BANK_DOWNLOAD_FILE, MONTH, TRANSFER_FILE


def run_transfers_update():
    current_month_number = datetime.now().month
    month_names = [
        "Enero",
        "Febrero",
        "Marzo",
        "Abril",
        "Mayo",
        "Junio",
        "Julio",
        "Agosto",
        "Septiembre",
        "Octubre",
        "Noviembre",
        "Diciembre",
    ]
    current_month_sheet = month_names[current_month_number - 1]
    print(f"   📂 Updating sheet: {current_month_sheet}")

    df_bank = pd.read_excel(BANK_DOWNLOAD_FILE, skiprows=1, engine="openpyxl")
    df_bank = df_bank.drop(columns=["Descripción"])
    df_bank.columns = df_bank.columns.str.strip()
    df_bank = df_bank.rename(
        columns={
            "Número Secuencia": "Nº Secuencia",
            "Descripción Extendida": "Descripción",
        }
    )

    expected_columns = ["Nº Secuencia", "Fecha", "Descripción", "Importe", "Saldo"]
    df_bank = df_bank[expected_columns]

    wb = load_workbook(TRANSFER_FILE)
    last_sequence = None

    if current_month_sheet not in wb.sheetnames:
        print(f"❌ Error: Sheet '{current_month_sheet}' not found in master file.")
        wb.create_sheet(MONTH)
        wb.save(TRANSFER_FILE)
        print(f"✅ {MONTH} sheet created successfully in {TRANSFER_FILE}")

    previous_sheet = wb[month_names[current_month_number - 2]]
    current_sheet = wb[current_month_sheet]
    print([cell.value for cell in current_sheet[1]])

    is_empty = all(
        all(cell.value is None for cell in row)
        for row in current_sheet.iter_rows(min_row=2, values_only=False)
    )

    if is_empty:
        header_row = list(
            previous_sheet.iter_rows(min_row=1, max_row=1, values_only=False)
        )[0]
        for col_idx, source_cell in enumerate(header_row, start=1):
            target_cell = current_sheet.cell(row=1, column=col_idx)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        for col_idx in range(1, previous_sheet.max_column + 1):
            col_letter = get_column_letter(col_idx)
            if col_letter in previous_sheet.column_dimensions:
                current_sheet.column_dimensions[
                    col_letter
                ].width = previous_sheet.column_dimensions[col_letter].width

        last_sequence = int(previous_sheet.cell(row=2, column=1).value)

    if last_sequence is None:
        seq_column = None
        header_row = list(
            current_sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
        for idx, col_name in enumerate(header_row):
            if col_name and "secuencia" in str(col_name).lower():
                seq_column = idx + 1
                break

        if not seq_column:
            print("❌ Could not find 'Nº Secuencia' column in master file.")
            return

        for row in current_sheet.iter_rows(min_row=2, values_only=True):
            if row[seq_column - 1]:
                last_sequence = int(row[seq_column - 1])
                break

        if last_sequence is None:
            print("❌ Could not determine last sequence number in master file.")
            return

    print(f"   📌 Last recorded sequence in master: {last_sequence}")

    df_new = df_bank[df_bank["Nº Secuencia"] > last_sequence]
    df_new = df_new.sort_values(by="Nº Secuencia", ascending=True)

    if df_new.empty:
        print("   ✅ No new transactions to add.")
        return

    print(f"📌 Adding {len(df_new)} new transactions to {current_month_sheet}.")

    for row in df_new.itertuples(index=False):
        current_sheet.insert_rows(2)
        for col_num, value in enumerate(row, start=1):
            current_sheet.cell(row=2, column=col_num, value=value)

    wb.save(TRANSFER_FILE)
    wb.close()
    print("   ✅ Transfer file updated successfully")


if __name__ == "__main__":
    run_transfers_update()
