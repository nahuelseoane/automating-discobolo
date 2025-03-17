import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import os
from dotenv import load_dotenv
from filter_payments import select_month

load_dotenv()

# Loading main variables
YEAR = os.getenv("YEAR")
TRANSFER_FILE = os.getenv("TRANSFER_FILE")
BASE_PATH = os.getenv("BASE_PATH")
MONTH_NUMBER = int(os.getenv("MONTH_NUMBER"))
MONTH = select_month(MONTH_NUMBER)
EMAILS_FILE = f"{BASE_PATH}/{YEAR}/EmailSocios.xlsx"
PAYMENT_PATH = f"{BASE_PATH}/{YEAR}/{MONTH_NUMBER} {MONTH} {YEAR}"
TRANSFER_FILE = f"{BASE_PATH}/{YEAR}/Transferencias {YEAR}.xlsx"
BANK_FILE = f"{BASE_PATH}/{YEAR}/MovimientosBanco.xlsx"
SHEET_NAME = MONTH
MOROSOS_DAILY = f"{BASE_PATH}/Morosos/descarga_reporte/reporte_morosos.xlsx"
MOROSOS_MAIN = f"{BASE_PATH}/Morosos/Morosos {YEAR}.xlsx"


df_daily = pd.read_excel(MOROSOS_DAILY, skiprows=5, engine="openpyxl")
wb = load_workbook(MOROSOS_MAIN)

# Checking if actual month sheet exists
try:
    with pd.ExcelFile(MOROSOS_MAIN, engine="openpyxl") as xls:
        sheet_names = xls.sheet_names  # Get all sheet names

    if MONTH not in sheet_names:
        print(f"📂 {MONTH} sheet not found. Creating it...")

        wb.create_sheet(MONTH)  # Create new sheet
        wb.save(MOROSOS_MAIN)

        # print(f"✅ {MONTH} sheet created successfully.")
    # else:
        # print(f"✅ {MONTH} sheet already exists.")

except FileNotFoundError:
    print("❌ Error: Main Morosos file not found!")


ws = wb[MONTH]
columns = [cell.value for cell in ws[1]]

# Modify excel
df_daily = df_daily.drop(columns=['NROCUENTA', 'FECHAULT.MOVDEBITO', 'MONTOULT.MOVDEBITO', 'DESC.ULT.MOVDEBITO',
                         'FECHAULT.MOVCREDITO', 'MONTOULT.MOVCREDITO', 'TELEFONO', 'EMAIL', 'SUCURSALEMISORA'])
df_daily = df_daily.sort_values(by=df_daily.columns[1], ascending=False)
df_daily = df_daily.drop(index=0)
# Reset index
df_daily = df_daily.iloc[1:].reset_index(drop=True)
# df_daily = df_daily.drop(index=0).reset_index(drop=True)
# print(f"📌 Morosos Daily columns: {df_daily.columns.tolist()}")

# Remove 0 or less values
df_daily = df_daily[df_daily.iloc[:, 1] > 0]

# Change 2do column name to 'SALDO'
df_daily.columns.values[1] = "SALDO"

# Open the writer and set the correct engine
with pd.ExcelWriter(MOROSOS_MAIN, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_daily.to_excel(writer, sheet_name=MONTH, index=False)

    # Access the workbook and the active worksheet
    workbook = writer.book
    worksheet = writer.sheets[MONTH]

    # ✔️ Adjusted styling
    header_fill = PatternFill(
        start_color='4DA6D2', end_color='4DA6D2', fill_type='solid')  # Excel-style blue
    header_font = Font(color='FFFFFF', bold=True)  # Bold white text

    # Apply the formatting to each header cell
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    # ✔️ Auto-adjust column widths
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = max_length + 2

    # ✔️ Format second column (balance) as currency
    for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = '"$"#,##0.00'

print("   ✅ Morosos main file updated.")
wb.close()
