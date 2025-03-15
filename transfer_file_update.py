import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# File paths
transfer_file = "${BASE_PATH}/${YEAR}/Transferencias ${YEAR}.xlsx"
bank_file = "${BASE_PATH}/${YEAR}/MovimientosBanco.xlsx"

# Get current month index (1 = January, 2 = February, ...)
current_month_number = datetime.now().month

month_names = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
               "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
current_month_sheet = month_names[current_month_number - 1]
print(f"📂 Updating sheet: {current_month_sheet}")

# Load bank file
df_bank = pd.read_excel(bank_file, skiprows=1)

# Delete "descripción" column (duplicate one)
df_bank = df_bank.drop(columns=["Descripción"])

# ✅ Fix column names for consistency
df_bank.columns = df_bank.columns.str.strip()
df_bank = df_bank.rename(columns={
    "Número Secuencia": "N° Secuencia",
    "Descripción Extendida": "Descripción"
})

# ✅ Ensure column order matches the master file
expected_columns = ["N° Secuencia", "Fecha", "Descripción", "Importe", "Saldo"]
df_bank = df_bank[expected_columns]

# ✅ Load the master file without modifying formatting
wb = load_workbook(transfer_file)
if current_month_sheet not in wb.sheetnames:
    print(f"❌ Error: Sheet '{current_month_sheet}' not found in master file.")
    exit()

ws = wb[current_month_sheet]

# ✅ Find the last recorded sequence in master file
seq_column = None
for col in ws.iter_cols(min_row=1, max_row=1, values_only=True):
    if "N° Secuencia" in col:
        seq_column = col.index("N° Secuencia") + 1
        break

if not seq_column:
    print("❌ Could not find 'N° Secuencia' column in master file.")
    exit()

last_sequence = None
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[seq_column - 1]:  # Ensure it's not empty
        last_sequence = int(row[seq_column - 1])
        break

if last_sequence is None:
    print("❌ Could not determine last sequence number in master file.")
    exit()
print(f"📌 Last recorded sequence in master: {last_sequence}")

# ✅ Filter only new transactions
df_new = df_bank[df_bank["N° Secuencia"] > last_sequence]
df_new = df_new.sort_values(by="N° Secuencia", ascending=True)  # Correct order

if df_new.empty:
    print("✅ No new transactions to add.")
    exit()

print(f"📌 Adding {len(df_new)} new transactions to {current_month_sheet}.")

# ✅ Insert new rows before existing transactions (to preserve format)
for row in df_new.itertuples(index=False):
    ws.insert_rows(2)  # Always insert at row 2 (below column headers)
    for col_num, value in enumerate(row, start=1):
        ws.cell(row=2, column=col_num, value=value)

# ✅ Save without losing formatting
wb.save(transfer_file)
wb.close()
print("   ✅ Transfer file updated successfully")
