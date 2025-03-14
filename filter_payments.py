import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def load_and_filter_payments(excel_file, sheet_name):
    """Load excel file and filter payments by 'Cuota' in the 'Concept' column."""
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        df_filtered = df[df["Concepto"].str.contains(
            "Cuota", case=False, na=False)].copy()

        df_filtered.reset_index(drop=True, inplace=True)
        print(
            f"✅ Loaded {len(df_filtered)} payments from {sheet_name} (Cuotas) only.")

        return df, df_filtered
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")


def filter_positive_payments(excel_file, sheet_name):
    """Load excel file and filter only positive payments."""
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        df_filtered = df[df['Importe'] > 0].copy()
        print(f"✅ Loadad {len(df_filtered)} payments from {sheet_name}.")
        return df, df_filtered
    except Exception as e:
        print(f"❌ Error loading Excel file: {e}")


def sanitize_filename(name):
    """Remove invalid characters for filename."""
    return re.sub(r'[\/:*?"<>|,]', '', name)


def update_loaded_status(excel_file, sheet_name, seq):
    """Mark the 'Sytech' column as 'Si' after a successful payment."""

    # Define gray fill style
    gray_fill = PatternFill(start_color="D3D3D3",
                            end_color="D3D3D3", fill_type="solid")
    try:
        wb = load_workbook(excel_file)

        if sheet_name not in wb.sheetnames:
            print(f"❌ Error: Sheet '{sheet_name}' not found in {excel_file}")
            return

        ws = wb[sheet_name]

        # ✅ Find the correct row for the client
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            if row[0].value == seq:  # Assuming is Column G
                sent_cell = row[7]  # Column H (Sytech)
                importe_cell = row[3]  # Column D (Importe)
                importe_cell.fill = gray_fill
                # ✅ Update the cell without affecting formatting
                sent_cell.value = "Si"

        # ✅ Save the file without overwriting sheets or formatting
        wb.save(excel_file)
        wb.close()
    except Exception as e:
        print(f"❌ Error updating load status: {e}")


def extract_operation_number(description):
    """Extracts the transaction number from 'Descripción'."""
    if not description:
        return None

    try:
        description = str(description)

        match = re.search(r'[CS]\.(\d+)', description)

        if match:
            return match.group(1)  # ✅ Extracted number or None

        numbers = re.findall(r"\d+", description)
        if numbers:
            return numbers[0]

        return None
    except Exception as e:
        print(f"❌ Error extracting operation number: {e}")


def extract_date(description):
    """Extracts the first occurrence of a date in DD/MM format from any transaction description."""
    match = re.search(r"(\d{2}/\d{2})",
                      description)  # Looks for first DD/MM pattern
    # Return the date if found, else None
    return match.group(1) + '/${YEAR}' if match else None


def extract_dni(description):
    """Extracts DNI from the given description.

    1. First, tries to find a pattern like 'C:12345678' or 'D:12345678'.
    2. If not found, searches for an 11-digit number that has nothing before or after.
    3. Converts the found number to an 8-digit DNI by removing the first two digits and the last digit.
    4. If no number was found, try searching for ORI:.
    """
    match = re.search(r'[CD]\:(\d+)', description)  # First pattern
    if match:
        return int(match.group(1)[2:10])  # Extract 8-digit DNI
    # Search for an 11-digit number with word boundaries (no letters or numbers before/after)
    match_2 = re.search(r'\b(\d{11})\b', description)
    if match_2:
        return int(match_2.group(1)[2:10])  # Extract 8-digit DNI
    match_3 = re.search(r'ORI:([0-9\-]+)', description)
    if match_3:
        return match_3.group(1)

    return None  # If no match found
