import pandas as pd
from openpyxl import load_workbook
from scripts.extra_functions import extract_dni, filter_positive_payments, extract_deposito
from config.config import TRANSFER_FILE, SHEET_NAME, EMAILS_FILE, TENNIS_CLASS_FEE

df, transfer = filter_positive_payments(TRANSFER_FILE, SHEET_NAME)
df_emails = pd.read_excel(EMAILS_FILE, sheet_name=SHEET_NAME)

transfer["DNI"] = transfer["Descripción"].apply(extract_dni)

# Merge
df_merged = transfer.merge(
    df_emails[['DNI', 'Jefe de Grupo I', 'Tipo de Pago']],
    on='DNI',
    how='left'
)

# print(df_merged[["Jefe de Grupo", "DNI"]].head())

# Load the workbook
wb = load_workbook(TRANSFER_FILE)
ws = wb[SHEET_NAME]

for idx, row in df_merged.iterrows():
    full_name = row['Jefe de Grupo I']
    description = row["Descripción"]
    seq, amount, tipo = row['N° Secuencia'], row['Importe'], row['Tipo de Pago']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        if row[0].value == seq:  # Assuming is Column A (N° Secuencia)
            jefe_de_grupo = row[6]  # Column H (Jefe de Grupo)
            concept = row[5]
            if concept.value is None or concept.value == "":  # Column 'Concept'
                if extract_deposito(description):
                    concept.value = 'DEPOSITO'
                    print(f" ✅ {seq} Updated--> Concept: DEPOSITO")
                elif not pd.isna(full_name):
                    if full_name == 'CANFORA, KEVIN' and int(amount) <= 20000:
                        concept.value = 'TENIS'
                        jefe_de_grupo.value = 'CLASES KEVIN'
                        print(
                            f" ✅ {seq} Updated --> Concept: TENIS & Jefe de Grupo: CLASES KEVIN")
                        continue

                    elif not pd.isna(tipo):
                        concept.value = str(tipo).strip()
                    else:
                        concept.value = 'CUOTA'
                    # Assigning Full name to column 'Jefe de Grupo'
                    if jefe_de_grupo.value is None or jefe_de_grupo.value == "":
                        jefe_de_grupo.value = full_name  # ✅ Update the cell without affecting formatting
                        if full_name:
                            print(
                                f" ✅ {seq} Updated --> Concept: {concept.value} & Jefe de Grupo: {full_name}.")

                    # print(f" ✅ {seq} - Concept updated to {concept.value}")
                elif int(amount) == TENNIS_CLASS_FEE:
                    concept.value = 'TENIS'
                    jefe_de_grupo.value = 'CLASE NO SOCIO'
                    print(
                        f" ✅ {seq} Updated --> Concept: TENIS & Jefe de Grupo: CLASE NO SOCIO")
                    continue
                else:
                    print(
                        f"  ❌ {seq} Not updated: Transfer's match wasn't found")

# Save the workbook as a new file
wb.save(TRANSFER_FILE)
wb.close()
print("   ✅ Updated 'Jefe de Grupo' and 'Concepto'")
