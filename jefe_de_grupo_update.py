import pandas as pd
from openpyxl import load_workbook
from extra_functions import extract_dni, filter_positive_payments
from config import TRANSFER_FILE, SHEET_NAME, EMAILS_FILE

df, transfer = filter_positive_payments(TRANSFER_FILE, SHEET_NAME)
df_emails = pd.read_excel(EMAILS_FILE, sheet_name=SHEET_NAME)

transfer["DNI"] = transfer["Descripción"].apply(extract_dni)

# Merge
df_merged = transfer.merge(
    df_emails[['DNI', 'Jefe de Grupo I', 'Tipo de Pago']],
    on='DNI',
    how='left'
)

# print(df_merged[["Jefe de Grupo", "DNI"]].head())

# Load the workbook
wb = load_workbook(TRANSFER_FILE)
ws = wb[SHEET_NAME]

for idx, row in df_merged.iterrows():
    full_name = row['Jefe de Grupo I']
    seq, amount, tipo = row['N° Secuencia'], row['Importe'], row['Tipo de Pago']
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        if row[0].value == seq:  # Assuming is Column A (N° Secuencia)
            sent_cell = row[6]  # Column H (Jefe de Grupo)
            if sent_cell.value is None or sent_cell.value == "":
                sent_cell.value = full_name  # ✅ Update the cell without affecting formatting
                print(f" ✅ {seq} Jefe de Grupo updated: {full_name}.")
                if row[5].value is None or sent_cell.value == "":  # Column 'Concept'
                    if not pd.isna(full_name):
                        if full_name == 'CANFORA, KEVIN' and int(amount) <= 20000:
                            row[5].value = 'TENIS'
                            row[6].value = 'CLASES KEVIN'
                            print(f" ✅ {seq} updated: TENIS - CLASES KEVIN")
                            continue

                        if not pd.isna(tipo):
                            row[5].value = str(tipo).strip()
                        else:
                            row[5].value = 'CUOTA'
                        print(f" ✅ Concept updated to {row[5].value} - {seq}")
                    if int(amount) == 5500:
                        row[5].value = 'TENIS'
                        row[6].value = 'CLASE NO SOCIO'
                        print(f" ✅ {seq} updated: TENIS - CLASE NO SOCIO")
                        continue


# Save the workbook as a new file
wb.save(TRANSFER_FILE)
wb.close()
print("   ✅ Updated 'Jefe de Grupo' and 'Concepto'")
